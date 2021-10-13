from itertools import chain, cycle
from typing import Iterable, Literal, Union

import pandas as pd
from numpy import Inf, NaN
from openpyxl import load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.table import Table
from pandas import DataFrame
from pandas.core.dtypes.common import is_list_like


class TableNotFound(Exception):
    pass


def table_to_df(
    ws: ReadOnlyWorksheet,
    table: Table,
    index,
    values_as_nan={"#NUM!", "#VALUE!", "#N/A", "#NAME?", "#REF!", "#NULL!"},
    values_as_inf={"#DIV/0!"},
    values_as_empty_string={None},
) -> pd.DataFrame:
    columns = [col.name for col in table.tableColumns]
    data_rows = ws[table.ref][
        (table.headerRowCount or 0) : -table.totalsRowCount
        if table.totalsRowCount is not None
        else None
    ]
    data = ((cell.value for cell in row) for row in data_rows)
    replacements = chain(
        zip(values_as_empty_string, cycle([""])),
        zip(values_as_nan, cycle([NaN])),
        zip(values_as_inf, cycle([Inf])),
    )
    frame = DataFrame(data, columns=columns, index=None)
    dtypes = frame.dtypes
    frame = frame.replace({k: r for k, r in replacements})
    for col, dtype in zip(frame.columns, dtypes):
        frame[col] = frame[col].astype(dtype)

    if index:
        if index == "auto":
            if table.tableStyleInfo.showFirstColumn:
                frame = frame.set_index(columns[0])
        elif index is False:
            pass
        elif is_list_like(index):
            frame = frame.set_index([columns[i] for i in index])
        else:
            frame = frame.set_index(columns[index])
    return frame


def xlsx_tables_to_dfs(
    file, index: Union[Literal["auto"], int, Iterable[int]] = "auto"
):
    """Get all tables from a given workbook. Returns a dictionary of tables.
    Requires a filename, which includes the file path and filename.

    Inspired by:
    https://github.com/pandas-dev/pandas/issues/24862#issuecomment-458885960
    https://stackoverflow.com/questions/43941365/openpyxl-read-tables-from-existing-data-book-example
    """

    # Load the workbook, from the filename, setting read_only to False
    wb = load_workbook(
        filename=file, read_only=False, keep_vba=False, data_only=True, keep_links=False
    )

    # Initialize the dictionary of tables
    return {
        name: table_to_df(ws, tbl, index)
        for ws in wb.worksheets
        for name, tbl in {**ws.tables}.items()
    }


def xlsx_table_to_df(
    file, table: str, index: Union[Literal["auto"], int, Iterable[int]] = "auto"
):
    """Get a table from a given workbook by the tablename."""

    # Load the workbook, from the filename, setting read_only to False
    wb = load_workbook(
        filename=file, read_only=False, keep_vba=False, data_only=True, keep_links=False
    )

    # Initialize the dictionary of tables

    for ws in wb.worksheets:
        if table in ws.tables:
            return table_to_df(ws, ws.tables[table], index)
    all_tables = {f"'{table}'" for ws in wb.worksheets for table in ws.tables.keys()}
    raise TableNotFound(
        f"Table '{table}' could not be found in the workbook. "
        f"Choose from {', '.join(all_tables)}."
    )
