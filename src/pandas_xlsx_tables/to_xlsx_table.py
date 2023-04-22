from typing import BinaryIO, Iterable, Literal, Optional, Tuple, Union

import numpy as np
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from pandas import DataFrame
import pandas as pd
import linecache
import sys,os
from .utils import NamedTableStyle, create_format_mapping, format_for_col
from typing import Iterable, Tuple, Optional, Union, BinaryIO
from typing import Optional, Union
from pandas.core.frame import DataFrame
import tempfile





HeaderOrientation = Literal["diagonal", "horizontal", "vertical"]


def check_datetime64(df):
    for col in df.columns:
        if pd.api.types.is_datetime64_ns_dtype(df[col].dtype):
            return True
    return False

def df_time_to_strings(df):
    """
    Check and convert datetime64[ns] columns to string format in a Pandas DataFrame.

    Args:
        df (Pandas DataFrame): The DataFrame to check and convert.

    Returns:
        Pandas DataFrame: The DataFrame with datetime columns converted to string format.
    """
    if check_datetime64(df): # check if there are datetime columns
        # create a copy of the input DataFrame
        df_copy = df.copy()
        # check for datetime columns and convert to string format using strftime()
        for col in df_copy.columns:
            if df_copy[col].dtype == 'datetime64[ns]':
                df_copy[col] = df_copy[col].dt.strftime('%Y-%m-%d %H:%M:%S')

        return df_copy
    else:
        return df


def copy_table(source_file, source_sheet, dest_file, dest_sheet):
    """
    Copy an Excel table from one workbook to another while preserving other sheets in the destination workbook.

    Args:
        source_file (str): The path to the source workbook.
        source_sheet (str): The name of the sheet containing the table in the source workbook.
        dest_file (str): The path to the destination workbook.
        dest_sheet (str): The name of the sheet where you want to copy the table to in the destination workbook.

    Returns:
        None
    """
    # Load the source workbook
    wb_source = load_workbook(source_file)

    

    # Load the destination workbook
    wb_dest = load_workbook(dest_file)
    
    # Get the sheet containing the table in the source workbook
    ws_source = wb_source[source_sheet]

    # Get the sheet where you want to copy the table to in the destination workbook
    try : ws_dest = wb_dest[dest_sheet]
    except: ws_dest = wb_dest.create_sheet(dest_sheet)

    # Copy the table from the source sheet to the destination sheet
    for row in ws_source.iter_rows(min_row=1, values_only=True):
        ws_dest.append(row)

    # Create a table in the destination sheet
    table = Table(displayName="Table1", ref=ws_dest.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws_dest.add_table(table)

    # Save the destination workbook
    wb_dest.save(dest_file)

def dfs_to_xlsx_tables(
    input: Iterable[Tuple[DataFrame, str]],
    file: Union[str, BinaryIO],
    index: bool = True,
    table_style: Optional[NamedTableStyle] = "Table Style Medium 9",
    nan_inf_to_errors=False,
    header_orientation: HeaderOrientation = "horizontal",
    remove_timezone: bool = False,
    append: bool = False,
) -> None:
    """Convert multiple dataframes to an excel file.

    Args:
        input (Iterable[Tuple[DataFrame, str]]): A list of tuples of (df, table_name)
        file (Union[str, BinaryIO]): File name or descriptor for the output
        index (bool, optional): Include the datafrme index in the results. Defaults
            to True
        table_style (Optional[NamedTableStyle], optional): Excel table style. Defaults
            to "Table Style Medium 9".
        nan_inf_to_errors (bool, optional): Explicitly write nan/inf values as errors.
            Defaults to False.
        header_orientation (HeaderOrientation, optional): Rotate the table headers, can
            be horizontal, vertical or diagonal. Defaults to "horizontal".
    """
    

    if append: #xlsxwriter does not support appending to existing files so we need to use openpyxl and tenp files as workaround
        #create a temp file with .xlsx extension

        
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx',delete=False) #create a temp file
        print ('tempfile',temp_file.name)
        temp_file_path = temp_file.name
        temp_file.close()
    else:
        temp_file_path = file
    print ('tempfilepath',temp_file_path)
    wb = xlsxwriter.Workbook(
        temp_file_path,
        options=dict(
            nan_inf_to_errors=nan_inf_to_errors,
            remove_timezone=remove_timezone,
        ),
    )

    format_mapping = create_format_mapping(wb)
    if header_orientation == "diagonal":
        header_format = wb.add_format()
        header_format.set_rotation(45)
    elif header_orientation == "vertical":
        header_format = wb.add_format()
        header_format.set_rotation(90)

    for df, table_name in input:
        ws = wb.add_worksheet(name=table_name)
        if index:
            df = df.reset_index()
        if not nan_inf_to_errors:
            df = (
                df.replace(np.Inf, np.finfo(np.float64).max)
                .replace(-np.Inf, np.finfo(np.float64).min)
                .fillna("")
            )

        column_names = (str(c) for c in df.columns)
        options = {
            "data": df.values,
            "name": table_name,
            "style": table_style,
            "first_column": index,
            "columns": [
                {"header": col_name, "format": format_for_col(df[col], format_mapping)}
                for col, col_name in zip(df.columns, column_names)
            ],
        }
        ws.add_table(0, 0, len(df), len(df.columns) - 1, options)

        if header_orientation == "diagonal":
            ws.set_row(
                0, max(15, 12 + 4 * max(len(c) for c in column_names)), header_format
            )
        elif header_orientation == "vertical":
            ws.set_row(
                0, max(15, 4 + 6 * max(len(c) for c in column_names)), header_format
            )
        elif header_orientation == "horizontal":
            # adjust row widths
            for i, width in enumerate(len(str(x)) for x in column_names):
                ws.set_column(i, i, max(8.43, width))
     
    wb.close()
    if append:
        for df, table_name in input:
            copy_table(source_file=temp_file_path, source_sheet=table_name, dest_file=file, dest_sheet=table_name)
        #copy_sheets_from_temp_to_original(temp_file_path, file)
        os.remove(temp_file_path)
    return



def PrintException():
    exc_type, exc_obj, tb = sys.exc_info()
    f = tb.tb_frame
    lineno = tb.tb_lineno
    filename = f.f_code.co_filename
    linecache.checkcache(filename)
    line = linecache.getline(filename, lineno, f.f_globals)
    print ( 'EXCEPTION IN ({}, LINE {} "{}"): {}'.format(filename, lineno, line.strip(), exc_obj))

def df_to_xlsx_table(
    df: DataFrame,
    table_name: str,
    file: Optional[Union[str, BinaryIO]] = None,
    index: bool = True,
    table_style: Optional[TableStyleInfo] = "Table Style Medium 9",
    nan_inf_to_errors=False,
    header_orientation: HeaderOrientation = "horizontal",
    remove_timezone: bool = False,
) -> None:
    """Convert single dataframe to an excel file.
    Deprecated: use df_to_xlsx_tables instead that works with single or multiple tables.
    Args:
        df (DataFrame): Padas dataframe to convert to excel.
        table_name (str):Name of the table.
        file (Union[str, BinaryIO]): File name or descriptor for the output.
            Defaults to <table_name>.xlsx
        index (bool, optional): Include the datafrme index in the results. Defaults
            to True
        table_style (Optional[NamedTableStyle], optional): Excel table style. Defaults
            to "Table Style Medium 9".
        nan_inf_to_errors (bool, optional): Explicitly write nan/inf values as errors.
            Defaults to False.
        header_orientation (HeaderOrientation, optional): Rotate the table headers, can
            be horizontal, vertical or diagonal. Defaults to "horizontal".
    """
    df=df_time_to_strings(df) #xlsxwriter does not support datetime64[ns] format so convert to strings if needed
    try : 
        dfs_to_xlsx_tables(
        [(df, table_name)],
        file=file or table_name + ".xlsx",
        index=index,
        table_style=table_style,
        nan_inf_to_errors=nan_inf_to_errors,
        header_orientation=header_orientation,
        remove_timezone=remove_timezone,
    )
    except Exception as e:
        print(e)
        print ("Error in df_to_xlsx_table")
        print ("If Permission denied - maybe file is open and needs to be closed first?")
        PrintException()
     


def df_to_xlsx_tables(tuple_or_list, file: Optional[Union[str, BinaryIO]] = None,
                   index: bool = True, table_style: Optional[Union[str, NamedTableStyle]] = "Table Style Medium 9",
                   nan_inf_to_errors: bool = False,
                   header_orientation: str = "horizontal",
                   remove_timezone: bool = False,
                   append=False) -> None:
                    
    """
    Export DataFrame(s) to XLSX file.

    Parameters:
    -----------
    tuple_or_list : tuple or list of tuples
        Tuple or list  containing either:
        - An iterable of (DataFrame, str) tuples to export multiple tables to a single workbook.
        - A single tuple (DataFrame, str) DataFrame and a table_name to export a single table to a workbook.

    file : Optional[Union[str, BinaryIO]], default None
        The file name or BinaryIO object to save the workbook.

    index : bool, default True
        Whether to include the DataFrame index in the exported table(s).

    table_style : Optional[Union[str, TableStyle]], default "Table Style Medium 9"
        The table style to apply to the exported table(s).

    nan_inf_to_errors : bool, default False
        Whether to convert NaN and infinite values to errors in the exported table(s).

    header_orientation : str, default "horizontal"
        The orientation of the table header in the exported table(s).

    remove_timezone : bool, default False
        Whether to remove the timezone from the datetime columns in the exported table(s).
    append : bool, default False

    Returns:
    --------
    None
    """
    print ('Saving xlsx table to ',file)
     
    if isinstance(tuple_or_list, Tuple) :
        # Export single table
        df, table_name = tuple_or_list
        df=df_time_to_strings(df) #xlsxwriter does not support datetime64[ns] format so convert to strings if needed
        
        data_list=[(df,table_name)]
    elif isinstance(tuple_or_list, Iterable):
        data_list2=[] #list of tuples with processed dataframes and table names
        data_list=tuple_or_list
        for df,table_name in data_list:
            df=df_time_to_strings(df)
            data_list2.append((df,table_name))
    else:
        print (tuple_or_list)
        raise ValueError("Invalid first argument passed to function.")
    try : 
        dfs_to_xlsx_tables(
            data_list2,
            file=file or table_name + ".xlsx",
            index=index,
            table_style=table_style,
            nan_inf_to_errors=nan_inf_to_errors,
            header_orientation=header_orientation,
            remove_timezone=remove_timezone,append=append
            )
       
    except Exception as e:
        print(e)
        print ("Error in df_to_xlsx_tables")
        print ("If Permission denied - maybe file is open and needs to be closed first?")
        PrintException()
