from typing import Iterable, Literal, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.table import Table
from pandas import DataFrame
from pandas.core.dtypes.common import is_list_like


def table_to_frame(ws: ReadOnlyWorksheet, table: Table, index) -> pd.DataFrame:
    columns = [col.name for col in table.tableColumns]
    data_rows = ws[table.ref][
        (table.headerRowCount or 0) : -table.totalsRowCount
        if table.totalsRowCount is not None
        else None
    ]
    data = ((cell.value for cell in row) for row in data_rows)

    frame = DataFrame(data, columns=columns, index=None)
    if index:
        if index == "auto":
            if table.tableStyleInfo.showFirstColumn:
                frame = frame.set_index(columns[0])
        elif is_list_like(index):
            frame = frame.set_index([columns[i] for i in index])
        else:
            frame = frame.set_index(columns[index])
    return frame


def xlsx_tables_to_frames(
    file, index: Union[Literal["auto"], int, Iterable[int]] = "auto"
):
    """Get all tables from a given workbook. Returns a dictionary of tables.
    Requires a filename, which includes the file path and filename.

    Inspired by:
    https://github.com/pandas-dev/pandas/issues/24862#issuecomment-458885960
    https://stackoverflow.com/questions/43941365/openpyxl-read-tables-from-existing-data-book-example
    """

    # Load the workbook, from the filename, setting read_only to False
    wb = load_workbook(
        filename=file, read_only=False, keep_vba=False, data_only=True, keep_links=False
    )

    # Initialize the dictionary of tables
    return {
        name: table_to_frame(ws, tbl, index)
        for ws in wb.worksheets
        for name, tbl in {**ws.tables}.items()
    }
